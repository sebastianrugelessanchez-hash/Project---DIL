import sys
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill

from path import get_report_file

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BATCH_FILL  = PatternFill("solid", fgColor="D6E4F0")
_BATCH_FONT  = Font(bold=True)
_FAIL_FILL   = PatternFill("solid", fgColor="FCE4D6")
_OK_FILL     = PatternFill("solid", fgColor="E2EFDA")  # verde claro para ya procesados


def _set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _write_batch_sheet(wb, sheet_name: str, col2_header: str, rows: list) -> None:
    """
    Crea una sheet con dos columnas (Ticket | <col2_header>) solo si hay filas.
    rows: lista de tuplas (ticket, doc_number).
    """
    if not rows:
        return
    ws = wb.create_sheet(sheet_name)
    ws.append(["Ticket", col2_header])
    for col in (1, 2):
        cell = ws.cell(1, col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for ticket, doc in rows:
        ws.append([ticket, doc])
        for col in (1, 2):
            ws.cell(ws.max_row, col).fill = _FAIL_FILL
    _set_col_widths(ws, [20, 30])


def write_report_xlsx(
    resultados: dict,
    total_tickets: int,
    vl06f: dict,
    zcmr_failures: dict,
    ticket_to_order: dict,
    cancel_failures: dict,
    tickets_no_encontrados: list = None,
    tickets_con_factura: list = None,
    tickets_critical_error: list = None,
    order_tracking: dict = None,
) -> None:
    """
    Genera el reporte del pipeline en xlsx con:
    - Sheet "Resumen": tabla con conteos por batch (procesados vs no encontrados)
    - Sheet "No encontrados en VL06F": tickets del Excel que VL06F no devolvió
    - Sheet "Con Factura": tickets excluidos por R001 (invoice_il empieza con '7')
    - Sheet "Error Crítico": tickets de chunks que crashearon (NO procesados)
    - Sheet "Tracking de Órdenes": lineage de cada ticket que tuvo orden pendiente
      en ZSD, con su estado en Batch 6, 7 y 8, y estado final consolidado.
    - Una sheet por batch fallido (solo si hay fallos), con Ticket + N° documento
    """
    tickets_no_encontrados = tickets_no_encontrados or []
    tickets_con_factura = tickets_con_factura or []
    tickets_critical_error = tickets_critical_error or []
    order_tracking = order_tracking or {}
    no_encontrados = len(tickets_no_encontrados)
    con_factura = len(tickets_con_factura)
    critical_err = len(tickets_critical_error)
    procesados = total_tickets - no_encontrados - con_factura - critical_err

    path = get_report_file()
    hoy  = date.today().strftime("%Y-%m-%d")

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet "Resumen" — conteos y tickets a revisar                       #
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Resumen"

    ws.append([f"REPORTE DIL — {hoy}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Tickets en Excel: {total_tickets}"])
    ws.append([f"Procesados por el pipeline: {procesados}"])
    if tickets_no_encontrados:
        ws.append([f"NO encontrados en VL06F: {no_encontrados} -> REQUIEREN INVESTIGACIÓN (ver sheet 'No encontrados en VL06F')"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="C00000")
    if tickets_con_factura:
        ws.append([f"Con factura intercompany (R001): {con_factura} -> EXCLUIDOS por tener factura ya emitida (ver sheet 'Con Factura')"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="9C5700")
    if tickets_critical_error:
        ws.append([f"Error crítico (chunk crasheó): {critical_err} -> NO procesados — RE-CORRER PIPELINE (ver sheet 'Error Crítico')"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="C00000")
    ws.append([])

    headers = ["Batch", "Exitosos", "Fallidos"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(ws.max_row, col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    tickets_con_fallo: set = set()
    for batch_name, (exitosos, fallidos) in resultados.items():
        ws.append([batch_name, len(exitosos), len(fallidos)])
        row = ws.max_row
        for col in (1, 2, 3):
            ws.cell(row, col).fill = _BATCH_FILL
            ws.cell(row, col).font = _BATCH_FONT
        if fallidos:
            ws.cell(row, 3).fill = _FAIL_FILL
        tickets_con_fallo.update(fallidos)

    ws.append([])
    completados = procesados - len(tickets_con_fallo)
    ws.append(["Exitosos (todos los batches):", f"{completados}/{procesados} ({completados}/{total_tickets} del Excel)"])
    ws.cell(ws.max_row, 1).font = _BATCH_FONT

    if tickets_con_fallo:
        ws.append(["Con fallos parciales:", f"{len(tickets_con_fallo)}/{procesados}"])
        ws.cell(ws.max_row, 1).font = _BATCH_FONT
        ws.append(["Revisar manualmente:", ", ".join(sorted(tickets_con_fallo))])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="C00000")

    if tickets_no_encontrados:
        ws.append(["NO encontrados en VL06F:", f"{no_encontrados}/{total_tickets} — pipeline no los detectó"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="C00000")

    if tickets_con_factura:
        ws.append(["Con factura intercompany (R001):", f"{con_factura}/{total_tickets} — excluidos del pipeline"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="9C5700")

    if tickets_critical_error:
        ws.append(["Error crítico (no procesados):", f"{critical_err}/{total_tickets} — chunk crasheó, RE-CORRER PIPELINE"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="C00000")

    _set_col_widths(ws, [40, 15, 15])

    # ------------------------------------------------------------------ #
    # Sheet "No encontrados en VL06F" — tickets que el pipeline no detectó #
    # ------------------------------------------------------------------ #
    if tickets_no_encontrados:
        ws_ne = wb.create_sheet("No encontrados en VL06F")
        ws_ne.append(["Ticket", "Motivo"])
        for col in (1, 2):
            cell = ws_ne.cell(1, col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for ticket in tickets_no_encontrados:
            ws_ne.append([ticket, "VL06F no devolvió este ticket — pipeline no lo procesó (requiere investigación)"])
            for col in (1, 2):
                ws_ne.cell(ws_ne.max_row, col).fill = _FAIL_FILL
        _set_col_widths(ws_ne, [20, 80])

    # ------------------------------------------------------------------ #
    # Sheet "Con Factura" — tickets excluidos por R001                     #
    # ------------------------------------------------------------------ #
    if tickets_con_factura:
        ws_cf = wb.create_sheet("Con Factura")
        ws_cf.append(["Ticket", "Invoice (ZZVBELN_IL)", "Motivo"])
        for col in (1, 2, 3):
            cell = ws_cf.cell(1, col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        _CF_FILL = PatternFill("solid", fgColor="FFE699")  # amarillo claro
        for ticket in tickets_con_factura:
            invoice = (vl06f.get(ticket, {}) or {}).get("invoice_il", "")
            ws_cf.append([ticket, invoice,
                          "R001: ya tiene factura intercompany — requiere cancelación manual de factura primero"])
            for col in (1, 2, 3):
                ws_cf.cell(ws_cf.max_row, col).fill = _CF_FILL
        _set_col_widths(ws_cf, [20, 25, 90])

    # ------------------------------------------------------------------ #
    # Sheet "Error Crítico" — tickets de chunks que crashearon            #
    # ------------------------------------------------------------------ #
    if tickets_critical_error:
        ws_ce = wb.create_sheet("Error Crítico")
        ws_ce.append(["Ticket", "Motivo", "Acción requerida"])
        for col in (1, 2, 3):
            cell = ws_ce.cell(1, col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for ticket in tickets_critical_error:
            ws_ce.append([
                ticket,
                "El chunk que contenía este ticket falló con excepción no controlada (ej: control SAP no encontrado, popup colgado)",
                "Re-correr 'python main.py' — el chunk se reintentará desde cero",
            ])
            for col in (1, 2, 3):
                ws_ce.cell(ws_ce.max_row, col).fill = _FAIL_FILL
        _set_col_widths(ws_ce, [20, 80, 60])

    # ------------------------------------------------------------------ #
    # Sheet "Tracking de Órdenes" — lineage ticket->order y estado por batch #
    # ------------------------------------------------------------------ #
    if order_tracking:
        ws_tr = wb.create_sheet("Tracking de Órdenes")
        headers_tr = ["Ticket", "Order", "Tipo", "Batch 6", "Batch 7", "Batch 8",
                      "Final Status", "Error Msg"]
        ws_tr.append(headers_tr)
        for col in range(1, len(headers_tr) + 1):
            cell = ws_tr.cell(1, col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

        # Fills por estado final para resaltado visual
        _STATUS_FILLS = {
            "deleted": PatternFill("solid", fgColor="E2EFDA"),  # verde claro
            "pending": PatternFill("solid", fgColor="FFE699"),  # amarillo
            "failed":  PatternFill("solid", fgColor="FCE4D6"),  # rojo claro
            "unknown": PatternFill("solid", fgColor="F2F2F2"),  # gris
        }

        for ticket in sorted(order_tracking.keys()):
            tr = order_tracking[ticket]
            tipo = "Intracompany" if tr.get("intracompany") else "Intercompany"
            tx = tr.get("transaction") or ("ME22N" if tr.get("intracompany") else "VA02")
            tipo_str = f"{tipo} ({tx})"
            final_status = tr.get("final_status", "unknown") or "unknown"
            ws_tr.append([
                ticket,
                tr.get("order", ""),
                tipo_str,
                tr.get("batch_6", "") or "-",
                tr.get("batch_7_verify", "") or "-",
                tr.get("batch_8_cancel", "") or "-",
                final_status,
                tr.get("error_msg", "") or "",
            ])
            fill = _STATUS_FILLS.get(final_status, _STATUS_FILLS["unknown"])
            for col in range(1, len(headers_tr) + 1):
                ws_tr.cell(ws_tr.max_row, col).fill = fill
        _set_col_widths(ws_tr, [16, 16, 22, 12, 12, 12, 16, 60])

    # ------------------------------------------------------------------ #
    # Sheets por batch fallido                                            #
    # ------------------------------------------------------------------ #
    def _fallidos(batch_key: str) -> list:
        return resultados.get(batch_key, ([], []))[1]

    def _vl06f_doc(t: str, field: str) -> str:
        return (vl06f.get(t, {}) or {}).get(field, "") or ""

    _write_batch_sheet(wb, "Billing Documents", "N° Billing Document",
        [(t, _vl06f_doc(t, "billing_doc")) for t in _fallidos("BATCH 1 — Billing Documents")])

    _write_batch_sheet(wb, "Shipment Cost", "N° Shipment Cost",
        [(t, _vl06f_doc(t, "shpt_cst")) for t in _fallidos("BATCH 2 — Shipment Cost")])

    _write_batch_sheet(wb, "Shipment Number", "N° Shipment Number",
        [(t, _vl06f_doc(t, "shipment")) for t in _fallidos("BATCH 3 — Shipment Number")])

    _write_batch_sheet(wb, "Reverse PGI", "N° Delivery",
        [(t, _vl06f_doc(t, "delivery")) for t in _fallidos("BATCH 4 — Reverse PGI")])

    _write_batch_sheet(wb, "BOL Deletion", "N° Delivery",
        [(t, _vl06f_doc(t, "delivery")) for t in _fallidos("BATCH 5 — BOL Deletion")])

    # Batch 6 — un ticket puede tener múltiples orders que fallaron
    zcmr_rows = [(t, ", ".join(orders)) for t, orders in zcmr_failures.items()]
    _write_batch_sheet(wb, "ZCMR Orders", "N° Order", zcmr_rows)

    # Batch 7 — tickets con orders pendientes según ZSD
    _write_batch_sheet(wb, "Verificación Final ZSD", "N° Order pendiente",
        [(t, ticket_to_order.get(t, "")) for t in _fallidos("BATCH 7 — Verificación Final ZSD")])

    # Batch 8 — orders que fallaron al cancelarse en VA02
    _write_batch_sheet(wb, "Order Cancellation", "N° Order",
        [(t, o) for t, o in cancel_failures.items()])

    # Guardar
    try:
        wb.save(path)
        print(f"  Reporte xlsx guardado en: {path}")
    except Exception as e:
        print(f"  [Reporte] No se pudo guardar el xlsx: {e}", file=sys.stderr)
