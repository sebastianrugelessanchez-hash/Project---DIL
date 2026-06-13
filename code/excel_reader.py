import sys
from pathlib import Path

import openpyxl

from path import get_billing_file
from sap_utils import _normalize_ticket

# Fila donde está el encabezado de las columnas en el Excel
_HEADER_ROW = 3

# Columnas que contienen ticket numbers. AMBAS son fuente válida — se mergea
# el contenido en una sola lista deduplicada. Si una fila tiene ZCMR vacía,
# normalmente el valor está en Ticket Number (y viceversa).
_ZCMR_COL = 3            # Columna C — "ZCMR"
_TICKET_NUMBER_COL = 10  # Columna J — "Ticket Number"

# Header esperado en cada columna (uppercase, sin spaces extra para comparar)
_HEADERS_EXPECTED = {
    _ZCMR_COL:           "ZCMR",
    _TICKET_NUMBER_COL:  "TICKET NUMBER",
}

# --- Manual Orders sheet (modo --manual-only) -------------------------------
# Hoja secundaria del mismo Excel con pares (ticket, order) para recovery
# de orders huérfanas que no se pudieron borrar en una corrida previa.
_MANUAL_SHEET_NAME = "Manual Orders"
_MANUAL_HEADER_ROW = 1
_MANUAL_TICKET_COL = 1  # Columna A
_MANUAL_ORDER_COL  = 2  # Columna B


def _coerce_to_ticket(val) -> str:
    """Normaliza un valor de celda a un ticket válido o '' si no aplica."""
    if val is None:
        return ""
    # Excel puede leer enteros como float (e.g. 70811039.0)
    if isinstance(val, float):
        if not val.is_integer():
            return ""
        val = int(val)
    ticket = _normalize_ticket(val)
    if not ticket or not ticket.isdigit():
        return ""
    return ticket


def _coerce_to_order(val) -> str:
    """
    Normaliza un valor de celda a un sales order SAP válido o '' si no aplica.
    Sales order: numérico, >=8 dígitos (típico SAP: 10 dígitos como 1150403358).
    """
    if val is None:
        return ""
    if isinstance(val, float):
        if not val.is_integer():
            return ""
        val = int(val)
    s = str(val).strip()
    if not s.isdigit() or len(s) < 8:
        return ""
    return s


def read_zcmr(filepath: Path = None) -> list[str]:
    """
    Lee tickets desde el Excel de billing combinando DOS columnas:
      - Columna C (ZCMR)
      - Columna J (Ticket Number)

    Ambas representan ticket numbers; cuando una está vacía la otra suele
    tener el valor. El pipeline procesa la unión (sin duplicados, preservando
    el orden en que aparecen).

    - Si no se pasa `filepath`, usa get_billing_file().
    - Omite filas/celdas vacías o no numéricas.
    - El color rojo del font en "Ticket Number" es metadata visual para el
      operador (indica "solo eliminar, no recargar") — el pipeline lo IGNORA.

    Lanza FileNotFoundError si no encuentra el archivo.
    Lanza ValueError si los encabezados de C3 o J3 no son los esperados.
    """
    path = filepath or get_billing_file()
    print(f"Leyendo tickets desde: {path.name}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    # Validar contrato del esquema: ambos encabezados deben coincidir
    for col, expected in _HEADERS_EXPECTED.items():
        actual = ws.cell(_HEADER_ROW, col).value
        actual_norm = str(actual or "").strip().upper()
        if expected not in actual_norm:
            wb.close()
            col_letter = openpyxl.utils.get_column_letter(col)
            raise ValueError(
                f"Se esperaba '{expected}' en fila {_HEADER_ROW} columna {col_letter}, "
                f"pero se encontró: '{actual}'. Verifica el formato del archivo."
            )

    tickets: list[str] = []
    seen: set = set()
    count_zcmr = 0
    count_tnum = 0

    # Leer ambas columnas en una sola pasada (más eficiente que dos iter_rows).
    # min_col=C, max_col=J trae 8 celdas por fila; usamos índices relativos.
    span_offset_zcmr = _ZCMR_COL - _ZCMR_COL            # 0
    span_offset_tnum = _TICKET_NUMBER_COL - _ZCMR_COL   # 7

    for row in ws.iter_rows(
        min_row=_HEADER_ROW + 1,
        min_col=_ZCMR_COL,
        max_col=_TICKET_NUMBER_COL,
        values_only=True,
    ):
        zcmr_val = row[span_offset_zcmr] if span_offset_zcmr < len(row) else None
        tnum_val = row[span_offset_tnum] if span_offset_tnum < len(row) else None

        for source, val in (("ZCMR", zcmr_val), ("Ticket Number", tnum_val)):
            ticket = _coerce_to_ticket(val)
            if not ticket or ticket in seen:
                continue
            seen.add(ticket)
            tickets.append(ticket)
            if source == "ZCMR":
                count_zcmr += 1
            else:
                count_tnum += 1

    wb.close()

    print(f"  {len(tickets)} tickets únicos cargados "
          f"(ZCMR={count_zcmr}, Ticket Number={count_tnum})")
    return tickets


def read_manual_orders(filepath: Path = None) -> dict:
    """
    Lee la hoja 'Manual Orders' del Excel: pares (Ticket, Order) para el
    modo --manual-only (recovery de orders huérfanas).

    Estructura esperada de la hoja:
        Fila 1: encabezados (Ticket | Order)
        Fila 2+: pares de valores

    - Retorna dict {ticket: order}.
    - Si la hoja NO existe -> retorna {} (no es error, solo modo normal).
    - Omite filas con celdas vacías o valores no numéricos.
    - Dedup por ticket (si el mismo ticket aparece dos veces, gana el último).

    Args:
        filepath: ruta al Excel. Si None, usa get_billing_file().
    """
    path = filepath or get_billing_file()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if _MANUAL_SHEET_NAME not in wb.sheetnames:
        available = list(wb.sheetnames)
        wb.close()
        # Diagnóstico: el match es exacto y case-sensitive. Si la hoja tiene
        # otro nombre/espacios/mayúsculas, aquí se ve cuál es el real.
        print(
            f"  [Manual Orders] Hoja '{_MANUAL_SHEET_NAME}' NO encontrada. "
            f"Hojas disponibles en el Excel: {available}",
            file=sys.stderr,
        )
        return {}

    ws = wb[_MANUAL_SHEET_NAME]
    pairs: dict = {}
    skipped_invalid = 0

    for row in ws.iter_rows(
        min_row=_MANUAL_HEADER_ROW + 1,
        min_col=_MANUAL_TICKET_COL,
        max_col=_MANUAL_ORDER_COL,
        values_only=True,
    ):
        if not row or len(row) < 2:
            continue
        ticket = _coerce_to_ticket(row[0])
        order = _coerce_to_order(row[1])

        if not ticket or not order:
            if row[0] is not None or row[1] is not None:
                skipped_invalid += 1
            continue

        pairs[ticket] = order

    wb.close()

    print(f"  [Manual Orders] {len(pairs)} pares (ticket, order) cargados "
          f"desde hoja '{_MANUAL_SHEET_NAME}'"
          + (f" (omitidos {skipped_invalid} por valores inválidos)" if skipped_invalid else ""))
    return pairs
